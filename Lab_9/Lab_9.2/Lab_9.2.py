import openpyxl

def average(values): # функцию для вычисления средних значений
    total = 0
    for value in values:
        total += value
    return round(total / len(values), 2)

wb = openpyxl.load_workbook("spreadsheet1.xlsx") # открываем таблицу
sheet = wb.active # берем текущий лист (т.е. лист 1)
workers = [sheet['A2':'I2'][0], sheet['A3':'I3'][0], sheet['A5':'I5'][0], sheet['A6':'I6'][0], sheet['A7':'I7'][0], sheet['A8':'I8'][0], sheet['A10':'I10'][0]] # выделяем из таблицы сотрудников
    
nomination = ["Человек с максимальной зарплатой:", "Человек с минимальной зарплатой:", "Средняя зарплата в бухгалтерии:","Средняя зарплата в отделе кадров:", "Средняя зарплата в столовой:"]
for i in range(0,5): # заполняем ячейки элементами из списка nomination
    cell = sheet.cell(row=14 + i, column=2)
    cell.value = nomination[0]
    del nomination[0]

maxSalaryWorker = max(workers, key=lambda x:x[8].value) # сотрудник с максимальной зарплатой
maxSalaryWorkerCell = sheet.cell(row=14, column=3); maxSalaryValueCell = sheet.cell(row=14,column=4) # обозначаем ячейки, в которые будет записана информация (сотрудник с макс. зп. и сама зп.)
maxSalaryWorkerCell.value = maxSalaryWorker[1].value; maxSalaryValueCell.value = maxSalaryWorker[8].value # записываем ифнормацию в эти колонки (сотрудник с макс. зп. и сама зп.)
    
minSalaryWorker = min(workers, key=lambda x:x[8].value) # сотрудник с минимальной зарплатой
minSalaryWorkerCell = sheet.cell(row=15, column=3); minSalaryValueCell = sheet.cell(row=15, column=4) # обозначаем ячейки, в которые будет записана информация (сотрудник с мин. зп. и сама зп.)
minSalaryWorkerCell.value = minSalaryWorker[1].value; minSalaryValueCell.value = minSalaryWorker[8].value # записываем информацию в эти колонки (сотрудник с мин. зп. и сама зп.)

averageDepartment1 = average([workers[0][8].value, workers[1][8].value]) # средняя зп в отделе бухгалтерии
averageDepartment2 = average([workers[2][8].value, workers[3][8].value, workers[4][8].value, workers[5][8].value]) # средняя зп в отделек адров
averageDepartment3 = average([workers[6][8].value]) # средняя зп в столовой
valueArray = [averageDepartment1, averageDepartment2, averageDepartment3] # добавляем средние значения в массив
for i in range(0,3): # заполняем соответствующие ячейки средних значений
    cell = sheet.cell(row=16 + i, column=3)
    cell.value = valueArray[i]

wb.save("spreadsheet2.xlsx") # сохраняем таблицу
    