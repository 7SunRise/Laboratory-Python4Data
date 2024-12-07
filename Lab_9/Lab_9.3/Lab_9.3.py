import openpyxl
from openpyxl.chart import PieChart, Reference 

wb = openpyxl.load_workbook("spreadsheet2.xlsx") # загружаем таблицу, полученную в задании 2
sheet = wb.active # берем активный лист (т.е. лист 1)

# Для удобства, создадим справа от таблицы "мини-таблицу" с нужными нам значениями
sheet.cell(2, 11).value = "Бухгалтерия"; sheet.cell(2, 12).value = sheet.cell(4,9).value
sheet.cell(3, 11).value = "Отдел кадров"; sheet.cell(3, 12).value = sheet.cell(9,9).value
sheet.cell(4, 11).value = "Столовая"; sheet.cell(4, 12).value = sheet.cell(11,9).value

pie = PieChart() # Создаем круговую диаграмму
labels = Reference(sheet, min_col=11, min_row=2, max_row=4) # Наименования ячеек (категории)
data = Reference(sheet, min_col=12, min_row=2, max_row=4) # Значения ячеек

pie.add_data(data) # Добавляем значения ячеек
pie.set_categories(labels) # Добавляем категории
pie.title = "Зарплата по отделам" # Добавляем заголовок диаграммы
sheet.add_chart(pie, 'M1') # Создаем диаграмму с верхним левым углом в ячейке M1

wb.save("spreadsheet3.xlsx") # Сохраняем таблицу